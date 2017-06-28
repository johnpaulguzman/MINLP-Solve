import openpyxl

import config

class XLSXReader:
    none_index = lambda array: array.index(None) if None in array else None

    def write_row(self, row):
        for index, item in enumerate(row):
            self.current_sheet.cell(row=self.row_counter, column=index+1, value=item)
        self.row_counter += 1

    def write_data(self, rows_list):
        book = openpyxl.load_workbook(self.xlsx_path)
        self.current_sheet = book.get_sheet_by_name(config.derived_data_sheet)
        self.row_counter = 0
        for index,cell in enumerate(self.current_sheet["A"]):
            if cell.value: self.row_counter = index + 3
        for rows in rows_list: self.write_row(rows)
        book.save(self.xlsx_path)

    def __init__(self, xlsx_path):
        self.round_places = 16 - 1 # config.decimal_precision = 16
        self.xlsx_path = xlsx_path
        self.data_chunks = self.extract_data_chunks(openpyxl.load_workbook(self.xlsx_path, data_only=True))

    def extract_data_chunks(self, book, exclude=None):
        data, data_chunks, data_dump = [], [], []
        for w in book.worksheets[:exclude]:
            for row in w.values:
                data += [row]
            data += [(None,)]
        for d in data:
            if d[0] is None: # note
                if data_dump: # note
                    data_chunks += [data_dump]
                    data_dump = []
            else: data_dump += [d] # note
        return data_chunks

    def init_idx(self, data_chunk):
        index_data = {}
        for index, index_range, *_ in data_chunk[1:]:
            lb, ub, *_ = index_range.split(config.range_infix)
            index_data[index] = list(range(int(lb), int(ub)+1))
        return index_data

    def parse_dim_info(self, dims_dict, dim_info):
        if not dim_info: return 
        dim_index, dim_value, *_ = dim_info.split(config.assign_delimiter)
        dims_dict[dim_index] = dim_value

    def process_number(self, number):
        return round(float(number), self.round_places)

    def extract_idxParams(self):
        P = {}
        idx = {}
        for data_chunk in self.data_chunks:
            if data_chunk[0][0] == config.index_keyword: idx = self.init_idx(data_chunk)
            else:
                index, *appended_dims_info = data_chunk[0][0].split(config.info_delimiter)
                if not P.get(index): P[index] = {}
                dims_dict = {}
                for dim_info in appended_dims_info: self.parse_dim_info(dims_dict, dim_info)
                dims = self.get_dims(index)
                row_size = len(data_chunk[:XLSXReader.none_index(data_chunk)])  # gets rid of None rows
                col_size = len(data_chunk[-1][:XLSXReader.none_index(data_chunk[-1])])  # gets rid of None columns
                if row_size == 1: P[index] = self.process_number(data_chunk[0][1])
                for row in range(1, row_size):
                    self.parse_dim_info(dims_dict, data_chunk[row][0])
                    for col in range(1, col_size):
                        self.parse_dim_info(dims_dict, data_chunk[0][col])
                        cell_value = self.process_number(data_chunk[row][col])
                        index_values = tuple([int(dims_dict.get(dim)) for dim in dims])
                        P[index][index_values] = cell_value
        return (idx, P)

    def get_dims(self, name):
        splitted = name.split(config.dims_delimiter)
        return tuple(splitted[1]) if len(splitted) == 2 else []

    def get_idx(self, name, idx):
        dims = self.get_dims(name)
        return tuple([(idx.get(dim.lower())) for dim in dims])

